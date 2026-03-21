"""
KB부동산 시세 분석기 - Mac 웹 버전 v1
실행: python3 "mac 매전차트 v1.py"
브라우저: http://localhost:5050
"""

from flask import Flask, render_template, request, jsonify, Response, send_file
import os, json, re, threading, webbrowser, warnings, queue, time, glob
import pandas as pd
import numpy as np
import requests
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from datetime import datetime, timedelta
import plotly.graph_objects as go
import plotly.io as pio

warnings.filterwarnings('ignore')

BASE_DIR      = os.path.dirname(os.path.abspath(__file__))
UPLOAD_DIR    = os.path.join(BASE_DIR, 'uploads')
SETTINGS_FILE = os.path.join(BASE_DIR, 'settings_mac.json')
os.makedirs(UPLOAD_DIR, exist_ok=True)

app = Flask(__name__, template_folder=os.path.join(BASE_DIR, 'templates'))
app.secret_key = 'kb_land_mac_2025'

DEFAULT_SERVICE_KEY = (
    "Vs5lXsSo6iEI8no3pP%2FT0udWF9s7Cc8oP1SIWnEI5F4h6dKq92fLvnKmx"
    "koWGJxSeW2%2FSOLQECGxOJzWcjJEXQ%3D%3D"
)

YEARLY_DISPOSABLE_INCOME = {
    2004:1150.2, 2005:1201.0, 2006:1252.9, 2007:1327.2, 2008:1428.1,
    2009:1460.4, 2010:1545.3, 2011:1622.8, 2012:1683.8, 2013:1749.5,
    2014:1826.1, 2015:1937.4, 2016:1982.5, 2017:2037.8, 2018:2119.7,
    2019:2205.9, 2020:2280.0, 2021:2364.4, 2022:2494.5, 2023:2632.3,
    2024:2774.2, 2025:2885.2, 2026:3000.6,
}

# ── 설정 ──────────────────────────────────────────────────────────────────────
def load_settings():
    if os.path.exists(SETTINGS_FILE):
        with open(SETTINGS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {'lawdong_path': '', 'service_key': DEFAULT_SERVICE_KEY}

def save_settings(data):
    with open(SETTINGS_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

# ── 법정동 코드 ───────────────────────────────────────────────────────────────
_region_cache = None

def load_region_codes(lawdong_path):
    global _region_cache
    if _region_cache is not None:
        return _region_cache
    codes = {}
    if not lawdong_path or not os.path.exists(lawdong_path):
        return codes
    for enc in ['cp949', 'euc-kr', 'utf-8']:
        try:
            with open(lawdong_path, 'r', encoding=enc) as f:
                for line in f:
                    parts = line.strip().split('\t')
                    if len(parts) < 2:
                        continue
                    code, name = parts[0].strip(), parts[1].strip()
                    if any('폐지' in p for p in parts) or code.endswith('00000'):
                        continue
                    sg_code = code[:5]
                    names = name.split()
                    if len(names) < 2:
                        continue
                    sido = names[0]
                    rest = names[1:]
                    for i in range(len(rest)-1, -1, -1):
                        if any(rest[i].endswith(s) for s in ['동','읍','면','가']):
                            dong    = rest[i]
                            sigungu = ' '.join(rest[:i])
                            if sigungu:
                                codes[(sido, sigungu, dong)] = (code, sg_code)
                            break
            _region_cache = codes
            return codes
        except UnicodeDecodeError:
            continue
    return codes

def find_sigungu_code(codes, sido, sigungu, dong):
    if (sido, sigungu, dong) in codes:
        return codes[(sido, sigungu, dong)][1]
    for (s, sg, d), (_, sc) in codes.items():
        if d == dong and sigungu in sg:
            return sc
    return None

# ── KB 엑셀 파싱 (openpyxl, win32com 제거) ───────────────────────────────────
def parse_kb_excel(path, sale_types=None, lease_types=None):
    sale_types  = sale_types  or ['normal']
    lease_types = lease_types or ['normal']

    complex_info = {}
    try:
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        def cv(r, c):
            v = ws.cell(r, c).value
            return str(v).strip() if v else ''
        complex_info = {
            '단지명':       cv(4,2),
            '대표번지':     cv(5,2),
            '공급전용면적': cv(7,2),
            '타입':         cv(8,2),
            '세대수':       cv(9,2),
            '방욕실':       cv(10,2),
            '현관구조':     cv(11,2),
        }
        wb.close()
    except Exception as e:
        print(f"단지 정보 오류: {e}")

    # 주소 파싱
    addr  = complex_info.get('대표번지', '')
    parts = addr.split()
    sido = sigungu = dong = ''
    for i in range(len(parts)-1, -1, -1):
        if any(parts[i].endswith(s) for s in ['동','읍','면','가']):
            dong = parts[i]
            for j in range(i-1, -1, -1):
                if parts[j].endswith(('구','군')):
                    sigungu = parts[j]
                    if parts[0].endswith('도'):
                        for k in range(j-1, -1, -1):
                            if parts[k].endswith('시'):
                                sigungu = parts[k]+' '+sigungu; break
                        sido = parts[0]
                    else:
                        sido = parts[0]
                    break
                elif parts[j].endswith('시'):
                    sigungu = parts[j]; sido = parts[0]; break
            break

    # 전용면적
    target_area = None
    at = complex_info.get('공급전용면적','')
    if '/' in at:
        try:
            s2 = ''.join(c for c in at.split('/')[-1] if c.isdigit() or c=='.')
            target_area = int(float(s2)) if s2 else None
        except: pass

    # 세대수
    type_households = None
    hi = complex_info.get('세대수','')
    for pat in [r'총\s*(\d+)\s*세대', r'\(.*?(\d{4,})\s*세대\)',
                r'(\d+)\s*세대\s*/\s*(\d+)\s*세대', r'(\d+)\s*/\s*(\d+)', r'(\d+)\s*세대']:
        m = re.search(pat, hi)
        if m:
            type_households = max(int(g) for g in m.groups()) if len(m.groups())>1 else int(m.group(1))
            break

    # 가격 데이터
    sale_col_map  = {'low':1,'normal':2,'high':3}
    lease_col_map = {'low':4,'normal':5,'high':6}
    priority      = {'high':0,'normal':1,'low':2}
    s_types = sorted(sale_types,  key=lambda x: priority[x])
    l_types = sorted(lease_types, key=lambda x: priority[x])

    df = pd.DataFrame()
    try:
        raw = pd.read_excel(path, skiprows=14)
        df['date'] = pd.to_datetime(
            raw.iloc[:,0].astype(str).apply(
                lambda x: f"{x[:4]}-{x[4:6]}-01" if len(x)>=6 else x),
            errors='coerce')
        df = df.dropna(subset=['date'])
        for t in s_types:
            df[f'매매가_{t}'] = pd.to_numeric(raw.iloc[:len(df), sale_col_map[t]], errors='coerce').fillna(0)
        for t in l_types:
            df[f'전세가_{t}'] = pd.to_numeric(raw.iloc[:len(df), lease_col_map[t]], errors='coerce').fillna(0)
        df['매매가'] = df[f'매매가_{s_types[0]}']
        df['전세가'] = df[f'전세가_{l_types[0]}']
    except Exception as e:
        print(f"가격 데이터 오류: {e}")

    return dict(complex=complex_info,
                address=dict(sido=sido, sigungu=sigungu, dong=dong, target_area=target_area),
                type_households=type_households, df=df,
                sale_types=s_types, lease_types=l_types)

# ── 실거래가 API ──────────────────────────────────────────────────────────────
def fetch_trades(sigungu_code, dong, apt_name, target_area, service_key):
    results = []
    now = datetime.now()
    for i in range(20*12):
        ym = (now - timedelta(days=30*i)).strftime("%Y%m")
        url = (f"http://apis.data.go.kr/1613000/RTMSDataSvcAptTrade/getRTMSDataSvcAptTrade"
               f"?serviceKey={service_key}&LAWD_CD={sigungu_code}&DEAL_YMD={ym}&numOfRows=1000")
        try:
            r = requests.get(url, timeout=10)
            if r.status_code != 200: continue
            root = ET.fromstring(r.text)
            for item in root.findall('.//item'):
                if (item.findtext('umdNm','').strip() != dong or
                        item.findtext('aptNm','').strip() != apt_name): continue
                area = float(item.findtext('excluUseAr','0'))
                if target_area and abs(int(area)-target_area) > 1: continue
                try:
                    price = float(item.findtext('dealAmount','0').replace(',',''))
                    fr    = item.findtext('floor','0')
                    floor = int(fr) if fr and fr.strip().lstrip('-').isdigit() else 0
                    date  = datetime(int(item.findtext('dealYear','2000')),
                                     int(item.findtext('dealMonth','1')), 1)
                    results.append({'date': date.isoformat(), 'price': price, 'floor': floor})
                except: pass
        except Exception as e:
            print(f"매매 API 오류 {ym}: {e}")
    return results

def fetch_leases(sigungu_code, dong, apt_name, target_area, service_key):
    results = []
    now = datetime.now()
    for i in range(20*12):
        ym = (now - timedelta(days=30*i)).strftime("%Y%m")
        url = (f"http://apis.data.go.kr/1613000/RTMSDataSvcAptRent/getRTMSDataSvcAptRent"
               f"?serviceKey={service_key}&LAWD_CD={sigungu_code}&DEAL_YMD={ym}&numOfRows=1000")
        try:
            r = requests.get(url, timeout=10)
            if r.status_code != 200: continue
            root = ET.fromstring(r.text)
            for item in root.findall('.//item'):
                if (item.findtext('umdNm','').strip() != dong or
                        item.findtext('aptNm','').strip() != apt_name): continue
                area = float(item.findtext('excluUseAr','0'))
                if target_area and abs(int(area)-target_area) > 1: continue
                if item.findtext('monthlyRent','0') not in ('0',''): continue
                try:
                    price = float(item.findtext('deposit','0').replace(',',''))
                    fr    = item.findtext('floor','0')
                    floor = int(fr) if fr and fr.strip().lstrip('-').isdigit() else 0
                    date  = datetime(int(item.findtext('dealYear','2000')),
                                     int(item.findtext('dealMonth','1')), 1)
                    results.append({'date': date.isoformat(), 'price': price, 'floor': floor})
                except: pass
        except: pass
    return results

# ── 심리지수 ──────────────────────────────────────────────────────────────────
def load_sentiment(path):
    try:
        wb = load_workbook(path, data_only=True)
        ws = wb['7.매수매도']
        regions = []
        for col in range(2, ws.max_column+1, 3):
            name = ws.cell(2, col).value
            if name: regions.append({'name': name, 'sell': col, 'buy': col+1})
        data = {}
        for row in range(3, ws.max_row+1):
            dc = ws.cell(row, 1).value
            if not dc: continue
            d = dc if isinstance(dc, datetime) else datetime.strptime(str(dc),'%Y-%m-%d')
            ds = d.strftime('%Y-%m-%d')
            data[ds] = {}
            for rg in regions:
                sv = ws.cell(row, rg['sell']).value
                bv = ws.cell(row, rg['buy']).value
                data[ds][rg['name']] = {
                    'sell': float(sv) if sv is not None else None,
                    'buy':  float(bv) if bv is not None else None,
                }
        wb.close()
        return {'regions': [r['name'] for r in regions], 'data': data}
    except Exception as e:
        print(f"심리 오류: {e}"); return None

# ── Plotly 차트 ───────────────────────────────────────────────────────────────
def build_chart(df, apt_name, area, trades, leases, sentiment_data, sentiment_region,
                show_pir, show_disposable_pir, show_una_sentiment,
                show_real_trade, sale_types, lease_types, type_households):

    SC = {'low':'#74B9FF','normal':'#2E86DE','high':'#0652DD'}
    LC = {'low':'#FFA502','normal':'#FF6348','high':'#EE5A24'}
    LB = {'low':'하위평균','normal':'일반평균','high':'상위평균'}
    fig = go.Figure()

    for t in sale_types:
        col = f'매매가_{t}'
        if col in df.columns:
            fig.add_trace(go.Scatter(x=df['date'], y=df[col], mode='lines',
                name=f'매매가({LB[t]})', line=dict(color=SC[t], width=2.5),
                hovertemplate='<b>%{x|%Y-%m}</b><br>매매가: %{y:,.0f}만원<extra></extra>'))

    for t in lease_types:
        col = f'전세가_{t}'
        if col in df.columns:
            fig.add_trace(go.Scatter(x=df['date'], y=df[col], mode='lines',
                name=f'전세가({LB[t]})', line=dict(color=LC[t], width=2, dash='dot'),
                hovertemplate='<b>%{x|%Y-%m}</b><br>전세가: %{y:,.0f}만원<extra></extra>'))

    # 전세가율
    sc = f'매매가_{sale_types[0]}' if sale_types else '매매가'
    lc = f'전세가_{lease_types[0]}' if lease_types else '전세가'
    if sc in df.columns and lc in df.columns:
        vdf = df[(df[sc]>0)&(df[lc]>0)].copy()
        if not vdf.empty:
            vdf['rate'] = vdf[lc]/vdf[sc]*100
            fig.add_trace(go.Scatter(x=vdf['date'], y=vdf['rate'], mode='lines',
                name='전세가율(%)', yaxis='y2',
                line=dict(color='#00B894', width=1.5, dash='dashdot'),
                hovertemplate='<b>%{x|%Y-%m}</b><br>전세가율: %{y:.1f}%<extra></extra>'))

    # PIR
    if show_pir and show_disposable_pir and sc in df.columns:
        pdf = df.copy()
        pdf['year']   = pd.to_datetime(pdf['date']).dt.year
        pdf['income'] = pdf['year'].map(YEARLY_DISPOSABLE_INCOME)
        pdf = pdf[pdf['income'].notna() & (pdf[sc]>0)].copy()
        if not pdf.empty:
            pdf['PIR'] = pdf[sc]/pdf['income']
            ra, total = [], 0
            for idx, v in enumerate(pdf['PIR'].values):
                total += v; ra.append(total/(idx+1))
            pdf['PIR_avg'] = ra
            fig.add_trace(go.Scatter(x=pdf['date'], y=pdf['PIR'], mode='lines',
                name='PIR(가처분)', yaxis='y3',
                line=dict(color='#E91E63', width=2.5),
                hovertemplate='<b>%{x|%Y-%m}</b><br>PIR: %{y:.2f}<extra></extra>'))
            fig.add_trace(go.Scatter(x=pdf['date'], y=pdf['PIR_avg'], mode='lines',
                name='PIR 누적평균', yaxis='y3',
                line=dict(color='#E91E63', width=1.5, dash='dash'),
                hovertemplate='<b>%{x|%Y-%m}</b><br>PIR평균: %{y:.2f}<extra></extra>'))

    # 실거래가
    if show_real_trade and trades:
        for min_f, max_f, nm, clr, op in [
            (5,999,'실거래(5층↑)','#3498DB',0.85),
            (0,4,  '실거래(4층↓)','#95A5A6',0.65)]:
            sub = [t for t in trades if min_f <= t['floor'] <= max_f]
            if sub:
                fig.add_trace(go.Scatter(
                    x=[t['date'] for t in sub], y=[t['price'] for t in sub],
                    mode='markers', name=nm,
                    marker=dict(color=clr, size=6, opacity=op,
                                line=dict(color='white', width=0.5)),
                    customdata=[[t['floor']] for t in sub],
                    hovertemplate='<b>%{x|%Y-%m}</b><br>실거래: %{y:,.0f}만원<br>층: %{customdata[0]}층<extra></extra>'))

    # 전세 실거래
    if leases:
        fig.add_trace(go.Scatter(
            x=[l['date'] for l in leases], y=[l['price'] for l in leases],
            mode='markers', name='전세실거래',
            marker=dict(color='#FFB366', size=5, opacity=0.75, symbol='diamond',
                        line=dict(color='white', width=0.5)),
            customdata=[[l.get('floor','-')] for l in leases],
            hovertemplate='<b>%{x|%Y-%m}</b><br>전세실거래: %{y:,.0f}만원<br>층: %{customdata[0]}층<extra></extra>'))

    # 심리지수
    if show_una_sentiment and sentiment_data and sentiment_region:
        raw = sentiment_data.get('data', {})
        xs, ys = [], []
        for ds, regions in sorted(raw.items()):
            rd = regions.get(sentiment_region, {})
            sv, bv = rd.get('sell'), rd.get('buy')
            if sv is not None and bv is not None and (sv+bv)>0:
                xs.append(ds); ys.append(round(bv/(sv+bv)*100, 1))
        if xs:
            fig.add_trace(go.Scatter(x=xs, y=ys, mode='lines',
                name=f'매수우위({sentiment_region})', yaxis='y4',
                line=dict(color='#6C5CE7', width=1.5),
                hovertemplate='<b>%{x|%Y-%m}</b><br>매수우위: %{y:.1f}%<extra></extra>'))
            fig.add_hline(y=50, line_dash='dash', line_color='#6C5CE7',
                          opacity=0.3, annotation_text='매수=매도', yref='y4')

    fig.update_layout(
        title=dict(text=f'<b>{apt_name} {area}㎡ KB시세 분석</b>',
                   font=dict(size=18, color='#2C3E50')),
        paper_bgcolor='white', plot_bgcolor='#FAFAFA',
        hovermode='x unified',
        legend=dict(orientation='h', yanchor='bottom', y=1.02,
                    xanchor='right', x=1, font=dict(size=10)),
        xaxis=dict(title='날짜', showgrid=True, gridcolor='#E8E8E8',
                   rangeslider=dict(visible=True, thickness=0.05)),
        yaxis=dict(title='가격 (만원)', showgrid=True, gridcolor='#E8E8E8', tickformat=','),
        yaxis2=dict(title='전세가율(%)', overlaying='y', side='right',
                    range=[0,100], showgrid=False, ticksuffix='%'),
        yaxis3=dict(title='PIR', overlaying='y', side='right',
                    showgrid=False, position=0.97, anchor='free'),
        yaxis4=dict(title='매수우위(%)', overlaying='y', side='right',
                    range=[0,100], showgrid=False, position=0.99, anchor='free'),
        margin=dict(l=60, r=120, t=80, b=60), height=640,
    )
    return fig

def fig_to_json(fig):
    return json.loads(pio.to_json(fig))

def fig_to_html(fig, apt_name, area, stats):
    """단독 실행 가능한 HTML 파일 생성"""
    stats_html = ''
    if stats:
        def fmt만(v):
            if not v: return '-'
            억 = v // 10000
            천 = round((v % 10000) / 1000)
            return (f'{억}억{천}천' if 천 else f'{억}억') if 억 else f'{round(v/1000)}천'

        rows = []
        if stats.get('current'):    rows.append(('현재 매매가', fmt만(stats['current']), ''))
        if stats.get('max'):        rows.append(('역대 최고가', fmt만(stats['max']), stats.get('max_date','')))
        if stats.get('drop_pct') is not None: rows.append(('고점 대비 하락', f"{stats['drop_pct']}%", ''))
        if stats.get('cagr'):       rows.append(('연복리(CAGR)', f"{stats['cagr']}%", '전기간'))
        if stats.get('pir_current'):rows.append(('현재 PIR', str(stats['pir_current']), f"평균 {stats.get('pir_avg','-')}"))
        if stats.get('trade_count'):rows.append(('매매 실거래', f"{stats['trade_count']}건", f"최근1년 {stats.get('recent_trade_count',0)}건"))
        if stats.get('turnover'):   rows.append(('연간 회전율', f"{stats['turnover']}%", ''))

        cards = ''.join(f'''
          <div style="background:#f8f9fa;border-radius:8px;padding:10px 14px;text-align:center;min-width:110px">
            <div style="font-size:11px;color:#95a5a6;margin-bottom:3px">{lb}</div>
            <div style="font-size:18px;font-weight:700;color:#2c3e50">{val}</div>
            <div style="font-size:11px;color:#7f8c8d">{sub}</div>
          </div>''' for lb, val, sub in rows)

        stats_html = f'''
        <div style="background:white;border-radius:10px;box-shadow:0 1px 4px rgba(0,0,0,.08);
                    padding:16px;margin-bottom:14px">
          <div style="font-size:13px;font-weight:700;color:#5d6d7e;margin-bottom:12px">📈 주요 지표</div>
          <div style="display:flex;flex-wrap:wrap;gap:8px">{cards}</div>
        </div>'''

    chart_div = pio.to_html(fig, include_plotlyjs='cdn', full_html=False, config={
        'displayModeBar': True, 'displaylogo': False,
        'modeBarButtonsToRemove': ['select2d','lasso2d']
    })

    now_str = datetime.now().strftime('%Y-%m-%d %H:%M')
    return f'''<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>{apt_name} {area}㎡ KB시세 분석</title>
<style>
  body{{font-family:"Apple SD Gothic Neo","Noto Sans KR",sans-serif;background:#f0f2f5;
        color:#2c3e50;margin:0;padding:16px}}
  header{{background:linear-gradient(135deg,#1a3a5c,#2980b9);color:white;
          padding:14px 24px;border-radius:10px;margin-bottom:14px;
          display:flex;align-items:center;gap:10px}}
  header h1{{font-size:17px;font-weight:700;margin:0}}
  header span{{font-size:12px;opacity:.7;margin-left:auto}}
  .chart-card{{background:white;border-radius:10px;
               box-shadow:0 1px 4px rgba(0,0,0,.08);padding:16px}}
</style>
</head>
<body>
<header>
  <svg width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2">
    <path d="M3 9l9-7 9 7v11a2 2 0 01-2 2H5a2 2 0 01-2-2z"/>
    <polyline points="9 22 9 12 15 12 15 22"/>
  </svg>
  <h1>KB부동산 시세 분석 — {apt_name} {area}㎡</h1>
  <span>생성: {now_str}</span>
</header>
{stats_html}
<div class="chart-card">{chart_div}</div>
</body>
</html>'''

# ── 통계 ──────────────────────────────────────────────────────────────────────
def compute_stats(df, sale_col, trades, type_households):
    s = {}
    if sale_col not in df.columns or df.empty: return s
    vdf = df[df[sale_col]>0].copy()
    if vdf.empty: return s
    vdf['date_dt'] = pd.to_datetime(vdf['date'])
    latest = vdf.iloc[-1]
    mi, xi = vdf[sale_col].idxmin(), vdf[sale_col].idxmax()
    s['current']  = int(latest[sale_col])
    s['max']      = int(vdf.loc[xi, sale_col]); s['max_date'] = str(vdf.loc[xi,'date'])[:7]
    s['min']      = int(vdf.loc[mi, sale_col]); s['min_date'] = str(vdf.loc[mi,'date'])[:7]
    s['drop_pct'] = round((s['max']-s['current'])/s['max']*100,1) if s['max']>0 else 0
    first = vdf.iloc[0]
    yrs = (latest['date_dt']-first['date_dt']).days/365.25
    if yrs>0 and float(first[sale_col])>0:
        s['cagr'] = round((np.power(float(latest[sale_col])/float(first[sale_col]),1/yrs)-1)*100,2)
    vdf['year']   = vdf['date_dt'].dt.year
    vdf['income'] = vdf['year'].map(YEARLY_DISPOSABLE_INCOME)
    pdf = vdf[vdf['income'].notna()]
    if not pdf.empty:
        li = pdf.iloc[-1]['income']
        if li and li>0:
            s['pir_current'] = round(float(latest[sale_col])/li, 2)
            s['pir_avg']     = round((pdf[sale_col]/pdf['income']).mean(), 2)
    if trades:
        s['trade_count'] = len(trades)
        recent = [t for t in trades
                  if datetime.fromisoformat(t['date'])>=datetime.now()-timedelta(days=365)]
        s['recent_trade_count'] = len(recent)
        if type_households and type_households>0:
            s['turnover'] = round(len(recent)/type_households*100,1)
    return s

# ── Flask 라우트 ──────────────────────────────────────────────────────────────
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/settings', methods=['GET'])
def get_settings():
    s = load_settings()
    path = s.get('lawdong_path','')
    return jsonify({'lawdong_path':path, 'service_key':s.get('service_key',DEFAULT_SERVICE_KEY),
                    'lawdong_loaded': bool(path and os.path.exists(path))})

@app.route('/api/settings', methods=['POST'])
def update_settings():
    global _region_cache
    data = request.json; s = load_settings()
    if 'lawdong_path' in data: s['lawdong_path']=data['lawdong_path']; _region_cache=None
    if 'service_key'  in data: s['service_key']=data['service_key']
    save_settings(s); return jsonify({'ok':True})

@app.route('/api/upload-lawdong', methods=['POST'])
def upload_lawdong():
    global _region_cache
    if 'file' not in request.files: return jsonify({'error':'파일 없음'}), 400
    p = os.path.join(UPLOAD_DIR,'law-dong.txt')
    request.files['file'].save(p); _region_cache=None
    codes = load_region_codes(p)
    if not codes: return jsonify({'error':'읽기 실패 (인코딩 확인)'}), 400
    s = load_settings(); s['lawdong_path']=p; save_settings(s)
    return jsonify({'ok':True, 'count':len(codes)})

@app.route('/api/upload-kb', methods=['POST'])
def upload_kb():
    if 'file' not in request.files: return jsonify({'error':'파일 없음'}), 400
    sale_types  = request.form.get('sale_types','normal').split(',')
    lease_types = request.form.get('lease_types','normal').split(',')
    p = os.path.join(UPLOAD_DIR,'kb_latest.xlsx')
    request.files['file'].save(p)
    parsed = parse_kb_excel(p, sale_types, lease_types)
    df = parsed['df']
    df_json = []
    for _, row in df.iterrows():
        d = {'date': str(row['date'])[:10]}
        for c in df.columns:
            if c != 'date': d[c] = float(row[c]) if pd.notna(row[c]) else 0
        df_json.append(d)
    return jsonify({'ok':True, 'complex':parsed['complex'], 'address':parsed['address'],
                    'type_households':parsed['type_households'], 'df':df_json,
                    'sale_types':parsed['sale_types'], 'lease_types':parsed['lease_types']})

@app.route('/api/upload-sentiment', methods=['POST'])
def upload_sentiment():
    if 'file' not in request.files: return jsonify({'error':'파일 없음'}), 400
    p = os.path.join(UPLOAD_DIR,'sentiment.xlsx')
    request.files['file'].save(p)
    data = load_sentiment(p)
    if data is None: return jsonify({'error':'로드 실패 (7.매수매도 시트 확인)'}), 400
    return jsonify({'ok':True, 'regions':data['regions']})

@app.route('/api/sentiment-regions')
def sentiment_regions():
    p = os.path.join(UPLOAD_DIR,'sentiment.xlsx')
    if not os.path.exists(p): return jsonify({'regions':[]})
    data = load_sentiment(p)
    return jsonify({'regions': data['regions'] if data else []})

@app.route('/api/analyze', methods=['POST'])
def analyze():
    body = request.json
    if not body.get('df'): return jsonify({'error':'KB 데이터 없음'}), 400
    df = pd.DataFrame(body['df'])
    df['date'] = pd.to_datetime(df['date'])

    apt_name        = body.get('apt_name','')
    area            = body.get('area','')
    sale_types      = body.get('sale_types',['normal'])
    lease_types     = body.get('lease_types',['normal'])
    type_households = body.get('type_households')
    address         = body.get('address',{})
    sido            = address.get('sido','')
    sigungu         = address.get('sigungu','')
    dong            = address.get('dong','')
    target_area     = address.get('target_area')

    show_pir            = body.get('show_pir', True)
    show_disposable_pir = body.get('show_disposable_pir', True)
    show_una_sentiment  = body.get('show_una_sentiment', True)
    show_real_trade     = body.get('show_real_trade', True)
    show_lease_trade    = body.get('show_lease_trade', True)
    sentiment_region    = body.get('sentiment_region','')

    s = load_settings()
    service_key  = s.get('service_key', DEFAULT_SERVICE_KEY)
    lawdong_path = s.get('lawdong_path','')

    trades = []; leases = []
    sigungu_code = None

    if show_real_trade and sido and sigungu and dong and apt_name:
        codes = load_region_codes(lawdong_path)
        if codes:
            sigungu_code = find_sigungu_code(codes, sido, sigungu, dong)
        if sigungu_code:
            trades = fetch_trades(sigungu_code, dong, apt_name, target_area, service_key)
            if show_lease_trade:
                leases = fetch_leases(sigungu_code, dong, apt_name, target_area, service_key)

    sentiment_data = None
    sp = os.path.join(UPLOAD_DIR,'sentiment.xlsx')
    if show_una_sentiment and os.path.exists(sp):
        sentiment_data = load_sentiment(sp)

    sc = f'매매가_{sale_types[0]}' if sale_types else '매매가'
    fig = build_chart(df, apt_name, area, trades, leases, sentiment_data, sentiment_region,
                      show_pir, show_disposable_pir, show_una_sentiment,
                      show_real_trade, sale_types, lease_types, type_households)
    stats = compute_stats(df, sc, trades, type_households)

    return jsonify({'ok':True, 'chart':fig_to_json(fig), 'stats':stats,
                    'trades_count':len(trades), 'leases_count':len(leases),
                    'sigungu_code':sigungu_code})

@app.route('/api/save-html', methods=['POST'])
def save_html():
    """차트를 독립 실행 HTML 파일로 저장 후 다운로드"""
    from flask import send_file
    body = request.json
    if not body.get('df'): return jsonify({'error':'KB 데이터 없음'}), 400

    df = pd.DataFrame(body['df'])
    df['date'] = pd.to_datetime(df['date'])

    apt_name        = body.get('apt_name','단지')
    area            = body.get('area','')
    sale_types      = body.get('sale_types',['normal'])
    lease_types     = body.get('lease_types',['normal'])
    type_households = body.get('type_households')
    address         = body.get('address',{})
    sido            = address.get('sido','')
    sigungu         = address.get('sigungu','')
    dong            = address.get('dong','')
    target_area     = address.get('target_area')

    show_pir            = body.get('show_pir', True)
    show_disposable_pir = body.get('show_disposable_pir', True)
    show_una_sentiment  = body.get('show_una_sentiment', True)
    show_real_trade     = body.get('show_real_trade', True)
    show_lease_trade    = body.get('show_lease_trade', True)
    sentiment_region    = body.get('sentiment_region','')

    s = load_settings()
    service_key  = s.get('service_key', DEFAULT_SERVICE_KEY)
    lawdong_path = s.get('lawdong_path','')

    trades = []; leases = []
    sigungu_code = None
    if show_real_trade and sido and sigungu and dong and apt_name:
        codes = load_region_codes(lawdong_path)
        if codes:
            sigungu_code = find_sigungu_code(codes, sido, sigungu, dong)
        if sigungu_code:
            trades = fetch_trades(sigungu_code, dong, apt_name, target_area, service_key)
            if show_lease_trade:
                leases = fetch_leases(sigungu_code, dong, apt_name, target_area, service_key)

    sentiment_data = None
    sp = os.path.join(UPLOAD_DIR,'sentiment.xlsx')
    if show_una_sentiment and os.path.exists(sp):
        sentiment_data = load_sentiment(sp)

    sc = f'매매가_{sale_types[0]}' if sale_types else '매매가'
    fig   = build_chart(df, apt_name, area, trades, leases, sentiment_data, sentiment_region,
                        show_pir, show_disposable_pir, show_una_sentiment,
                        show_real_trade, sale_types, lease_types, type_households)
    stats = compute_stats(df, sc, trades, type_households)

    # HTML 파일 저장 (원본 r5.py와 같은 폴더)
    safe_name = apt_name.replace(' ','_').replace('/','_')
    filename  = f"{safe_name}_{area}m2_chart.html"
    save_path = os.path.join(BASE_DIR, filename)

    html_content = fig_to_html(fig, apt_name, area, stats)
    with open(save_path, 'w', encoding='utf-8') as f:
        f.write(html_content)

    return send_file(save_path, as_attachment=True, download_name=filename,
                     mimetype='text/html')

# ── KB 자동 다운로드 (Selenium) ───────────────────────────────────────────────
# 진행상황을 SSE로 스트리밍하기 위한 큐
_kb_progress_queues: dict = {}

def kb_auto_download(apt_name: str, target_area: str, q: queue.Queue):
    """Selenium으로 KB부동산 시세 엑셀 자동 다운로드"""
    def send(msg_type, **kw):
        q.put({'type': msg_type, **kw})

    download_dir = UPLOAD_DIR

    try:
        from selenium import webdriver
        from selenium.webdriver.chrome.service import Service
        from selenium.webdriver.chrome.options import Options
        from selenium.webdriver.common.by import By
        from selenium.webdriver.common.keys import Keys
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from webdriver_manager.chrome import ChromeDriverManager
    except ImportError as e:
        send('error', message=f'selenium 미설치: {e}'); return

    # ── ChromeDriver 설치 (캐시 있으면 빠름) ──────────────────────────────────
    send('progress', step=1, message='ChromeDriver 준비 중...')
    try:
        driver_path = ChromeDriverManager().install()
    except Exception as e:
        send('error', message=f'ChromeDriver 설치 실패: {e}'); return

    abs_dl = os.path.abspath(download_dir)
    prefs = {
        'download.default_directory':   abs_dl,
        'download.prompt_for_download': False,
        'download.directory_upgrade':   True,
        'safebrowsing.enabled':         False,
        'profile.default_content_setting_values.automatic_downloads': 1,
    }

    # 비headless 모드 — KB부동산은 headless 감지하여 차단함. 로컬 Mac에서는 창 표시가 안정적
    opts = Options()
    # headless 미사용: KB부동산이 headless 차단
    opts.add_argument('--window-size=1400,900')
    opts.add_argument('--disable-gpu')
    opts.add_argument('--no-sandbox')
    opts.add_argument('--disable-dev-shm-usage')
    opts.add_argument('--disable-blink-features=AutomationControlled')
    opts.add_argument('--disable-extensions')
    opts.add_argument('--disable-popup-blocking')
    opts.add_argument('--user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) '
                      'AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36')
    opts.add_experimental_option('prefs', prefs)
    opts.add_experimental_option('excludeSwitches', ['enable-automation'])
    opts.add_experimental_option('useAutomationExtension', False)
    opts.page_load_strategy = 'normal'

    send('progress', step=2, message='Chrome 실행 중...')
    try:
        service = Service(driver_path)
        driver  = webdriver.Chrome(service=service, options=opts)
        driver.implicitly_wait(3)
    except Exception as e:
        send('error', message=f'Chrome 실행 실패: {e}'); return

    # 자동화 감지 방지
    try:
        driver.execute_cdp_cmd('Page.addScriptToEvaluateOnNewDocument', {
            'source': "Object.defineProperty(navigator,'webdriver',{get:()=>undefined})"})
        driver.execute_cdp_cmd('Page.setDownloadBehavior',
                               {'behavior':'allow','downloadPath': abs_dl})
    except: pass

    before_files = set(os.listdir(download_dir))

    try:
        # ── KB부동산 접속 ──────────────────────────────────────────────────────
        send('progress', step=3, message='KB부동산 접속 중... (Chrome 창이 잠시 열립니다)')
        driver.set_page_load_timeout(60)
        driver.get('https://kbland.kr/map')
        time.sleep(5)   # JS 렌더링 대기 (비headless는 더 빠름)

        # 디버그: 페이지 제목 확인
        print(f"[KB] 페이지 제목: {driver.title}")

        # 팝업 닫기
        for sel in ['button.close','button.popup-close','[aria-label="닫기"]',
                    '.modal-close','button[class*="close"]']:
            try:
                for btn in driver.find_elements(By.CSS_SELECTOR, sel):
                    if btn.is_displayed():
                        driver.execute_script('arguments[0].click();', btn)
                        time.sleep(0.5)
            except: pass

        # ── 검색창 열기 ────────────────────────────────────────────────────────
        send('progress', step=4, message=f'"{apt_name}" 검색 중...')
        search_opened = False
        search_selectors = [
            'div.mapsearch-wrap button',
            '#app > div > div.mapsearch-wrap > div > button',
            'button.btn-land-search',
            'button[class*="search"]',
            '//button[contains(@class,"search")]',
        ]
        for sel in search_selectors:
            try:
                if sel.startswith('//'):
                    btn = WebDriverWait(driver, 5).until(
                        EC.presence_of_element_located((By.XPATH, sel)))
                else:
                    btn = WebDriverWait(driver, 5).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, sel)))
                driver.execute_script('arguments[0].click();', btn)
                print(f"[KB] 검색 버튼 클릭 성공: {sel}")
                time.sleep(2); search_opened = True; break
            except: pass

        if not search_opened:
            # 페이지 소스 저장 (디버그용)
            dbg = os.path.join(BASE_DIR, 'kb_debug.html')
            with open(dbg, 'w', encoding='utf-8') as f: f.write(driver.page_source)
            send('error', message='검색 버튼을 찾을 수 없습니다. kb_debug.html 확인')
            driver.quit(); return

        # 입력창 대기
        inp_selectors = [
            "input[placeholder*='단지']",
            "input[placeholder*='주소']",
            "input.form-control",
            "input[type='text']",
        ]
        inp = None
        for sel in inp_selectors:
            try:
                inp = WebDriverWait(driver, 8).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, sel)))
                break
            except: pass

        if not inp:
            send('error', message='검색 입력창을 찾을 수 없습니다.'); driver.quit(); return

        inp.clear(); inp.send_keys(apt_name); time.sleep(1.5)
        inp.send_keys(Keys.RETURN); time.sleep(4)

        # ── 검색 결과 처리 ─────────────────────────────────────────────────────
        # widthTypeSelect = 단지 페이지 로드됨
        direct = False
        try:
            WebDriverWait(driver, 6).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, 'div.widthTypeSelect')))
            direct = True
        except: pass

        if not direct:
            items_el = driver.find_elements(By.CSS_SELECTOR, 'div.item-search-poi')
            if not items_el:
                # 검색 결과 없음 - 디버그 저장
                dbg = os.path.join(BASE_DIR, 'kb_debug.html')
                with open(dbg, 'w', encoding='utf-8') as f: f.write(driver.page_source)
                send('error', message=f'"{apt_name}" 검색 결과 없음. kb_debug.html 확인')
                driver.quit(); return

            if len(items_el) == 1:
                try:
                    span = items_el[0].find_element(By.CSS_SELECTOR, 'span.search-poi')
                    driver.execute_script('arguments[0].click();', span)
                except:
                    driver.execute_script('arguments[0].click();', items_el[0])
                time.sleep(4)
            else:
                choices = []
                for el in items_el:
                    try:
                        name = el.find_element(By.CSS_SELECTOR, 'span.text').text.strip()
                        loc  = el.find_element(By.CSS_SELECTOR, 'span.date').text.strip()
                        choices.append({'name': name, 'loc': loc})
                    except: choices.append({'name':'?', 'loc':'?'})

                send('select', choices=choices, message='단지를 선택해주세요.')
                sel_idx = None
                for _ in range(180):   # 90초 대기
                    if not q.empty():
                        msg = q.get()
                        if msg.get('type') == 'select_result':
                            sel_idx = msg['index']; break
                    time.sleep(0.5)
                if sel_idx is None:
                    send('error', message='단지 선택 시간 초과'); driver.quit(); return

                items_el = driver.find_elements(By.CSS_SELECTOR, 'div.item-search-poi')
                try:
                    span = items_el[sel_idx].find_element(By.CSS_SELECTOR, 'span.search-poi')
                    driver.execute_script('arguments[0].click();', span)
                except:
                    driver.execute_script('arguments[0].click();', items_el[sel_idx])
                time.sleep(4)

        # ── 면적 선택 ──────────────────────────────────────────────────────────
        send('progress', step=5, message='면적 선택 중...')
        target_int = str(int(float(target_area))) if target_area else ''

        # widthTypeSelect 드롭다운 열기
        try:
            dd = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, 'div.widthTypeSelect')))
            driver.execute_script('arguments[0].scrollIntoView({block:"center"});', dd)
            time.sleep(1)
            driver.execute_script('arguments[0].click();', dd)
            time.sleep(2.5)
        except Exception as e:
            send('error', message=f'면적 드롭다운 열기 실패: {e}'); driver.quit(); return

        # 면적 행 파싱 (여러 셀렉터 시도)
        def get_area_rows():
            for sel in ['div.tbody-tr', 'tr.tbody-tr', 'div[class*="tbody"]']:
                rows = driver.find_elements(By.CSS_SELECTOR, sel)
                if rows: return rows
            return []

        def parse_area_from_row(row):
            for sel in ['span.tdbox span.tbarea-point em', 'em', 'td', 'span']:
                try:
                    for el in row.find_elements(By.CSS_SELECTOR, sel):
                        m = re.search(r'전용\s*([\d.]+)', el.text)
                        if m: return m.group(1)
                except: pass
            return None

        rows = get_area_rows()
        clicked = False
        area_choices = []

        for row in rows:
            a = parse_area_from_row(row)
            if a:
                area_choices.append(a)
                if target_int and a.split('.')[0] == target_int:
                    driver.execute_script('arguments[0].click();', row)
                    time.sleep(2); clicked = True; break

        if not clicked:
            if area_choices:
                send('select_area', choices=area_choices, message='면적을 선택해주세요.')
                sel_area = None
                for _ in range(180):
                    if not q.empty():
                        msg = q.get()
                        if msg.get('type') == 'select_area_result':
                            sel_area = msg['area']; break
                    time.sleep(0.5)
                if not sel_area:
                    send('error', message='면적 선택 시간 초과'); driver.quit(); return
                target_int = sel_area.split('.')[0]
                rows = get_area_rows()
                for row in rows:
                    a = parse_area_from_row(row)
                    if a and a.split('.')[0] == target_int:
                        driver.execute_script('arguments[0].click();', row)
                        time.sleep(2); clicked = True; break
            else:
                # 면적 정보를 아예 못 찾은 경우 - 첫 번째 행 선택
                rows = get_area_rows()
                if rows:
                    driver.execute_script('arguments[0].click();', rows[0])
                    time.sleep(2); clicked = True

        if not clicked:
            send('error', message='면적 선택 실패'); driver.quit(); return

        # ── STEP 6: KB시세 버튼 클릭 (팝업 열기) ────────────────────────────
        send('progress', step=6, message='KB시세 버튼 클릭 중...')
        time.sleep(3)

        def find_btn_by_text(keywords, timeout=8):
            """키워드 중 하나라도 포함된 버튼 반환"""
            deadline = time.time() + timeout
            while time.time() < deadline:
                btns = driver.find_elements(By.TAG_NAME, 'button')
                for b in btns:
                    try:
                        t = b.text.strip()
                        if t and any(kw in t for kw in keywords):
                            return b, t
                    except: pass
                time.sleep(0.5)
            return None, None

        def find_btn_by_css(selectors, timeout=5):
            for sel in selectors:
                try:
                    b = WebDriverWait(driver, timeout).until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, sel)))
                    return b
                except: pass
            return None

        # 현재 버튼 목록 디버그 출력
        print("[KB] 현재 버튼 목록:")
        for b in driver.find_elements(By.TAG_NAME, 'button'):
            try:
                t = b.text.strip()
                if t: print(f"  '{t}'")
            except: pass

        # KB시세 버튼 (첫 번째 클릭 — 팝업/드롭다운 열기)
        kb_btn, kb_btn_text = find_btn_by_text(['KB시세', 'KB 시세', '시세표'])
        if not kb_btn:
            kb_btn = find_btn_by_css([
                'button.btn-land-sqlinebx',
                'button[class*="sqlinebx"]',
                'div.f-row2-gap8 > button:nth-child(2)',
                '#시세 button:last-child',
            ])
            kb_btn_text = kb_btn.text.strip() if kb_btn else '?'

        if not kb_btn:
            dbg = os.path.join(BASE_DIR, 'kb_debug.html')
            with open(dbg, 'w', encoding='utf-8') as f: f.write(driver.page_source)
            send('error', message='KB시세 버튼 없음. kb_debug.html 확인')
            driver.quit(); return

        print(f"[KB] KB시세 버튼 클릭: '{kb_btn_text}'")
        driver.execute_script('arguments[0].scrollIntoView({block:"center"});', kb_btn)
        time.sleep(0.5)
        driver.execute_script('arguments[0].click();', kb_btn)
        time.sleep(2.5)   # 팝업/드롭다운 열릴 때까지 대기

        # ── STEP 7: 과거시세 다운로드 버튼 클릭 (실제 다운로드) ──────────────
        send('progress', step=7, message='과거시세 다운로드 클릭 중...')

        print("[KB] KB시세 클릭 후 버튼 목록:")
        for b in driver.find_elements(By.TAG_NAME, 'button'):
            try:
                t = b.text.strip()
                if t: print(f"  '{t}'")
            except: pass

        dl_btn, dl_btn_text = find_btn_by_text(['과거시세', '과거 시세', '다운로드'], timeout=8)
        if not dl_btn:
            dl_btn = find_btn_by_css([
                'button[class*="history"]',
                'div.layer button',
                'div.popup button',
                'div[class*="modal"] button',
            ])
            dl_btn_text = dl_btn.text.strip() if dl_btn else '?'

        if not dl_btn:
            dbg = os.path.join(BASE_DIR, 'kb_debug.html')
            with open(dbg, 'w', encoding='utf-8') as f: f.write(driver.page_source)
            send('error', message='과거시세 다운로드 버튼 없음. kb_debug.html 확인')
            driver.quit(); return

        print(f"[KB] 과거시세 버튼 클릭: '{dl_btn_text}'")
        driver.execute_script('arguments[0].scrollIntoView({block:"center"});', dl_btn)
        time.sleep(0.5)
        driver.execute_script('arguments[0].click();', dl_btn)
        send('progress', step=8, message='파일 다운로드 중... (최대 60초)')
        time.sleep(3)

        # ── 다운로드 완료 대기 ─────────────────────────────────────────────────
        saved_path = None
        for _ in range(60):   # 최대 60초 대기
            time.sleep(1)
            cur = set(os.listdir(download_dir))
            new_files = cur - before_files
            # .crdownload 제외, .xlsx 또는 임시파일 없는 것
            completed = [f for f in new_files
                         if not f.endswith('.crdownload') and not f.endswith('.tmp')
                         and f.endswith('.xlsx')]
            if completed:
                saved_path = os.path.join(download_dir, completed[0])
                print(f"[KB] 다운로드 완료: {completed[0]}")
                break

        driver.quit()

        if not saved_path:
            send('error', message='파일 다운로드 실패 (60초 초과)'); return

        # kb_latest.xlsx로 이동
        import shutil
        dest = os.path.join(download_dir, 'kb_latest.xlsx')
        shutil.copy2(saved_path, dest)

        send('progress', step=9, message='엑셀 파싱 중...')
        parsed = parse_kb_excel(dest)
        df = parsed['df']
        df_json = []
        for _, row in df.iterrows():
            d = {'date': str(row['date'])[:10]}
            for c in df.columns:
                if c != 'date': d[c] = float(row[c]) if pd.notna(row[c]) else 0
            df_json.append(d)

        send('result', complex=parsed['complex'], address=parsed['address'],
             type_households=parsed['type_households'], df=df_json,
             sale_types=parsed['sale_types'], lease_types=parsed['lease_types'])

    except Exception as e:
        import traceback; traceback.print_exc()
        try: driver.quit()
        except: pass
        send('error', message=str(e))


@app.route('/api/kb-search-stream')
def kb_search_stream():
    """SSE 엔드포인트 — KB 자동 다운로드 진행상황 스트리밍"""
    apt_name    = request.args.get('apt_name','').strip()
    target_area = request.args.get('area','').strip()
    session_id  = request.args.get('sid','')

    if not apt_name:
        return jsonify({'error':'단지명 필요'}), 400

    q = queue.Queue()
    _kb_progress_queues[session_id] = q

    # 별도 스레드에서 Selenium 실행
    t = threading.Thread(target=kb_auto_download, args=(apt_name, target_area, q), daemon=True)
    t.start()

    def event_stream():
        while True:
            try:
                msg = q.get(timeout=180)   # 3분 대기 (비headless 모드에서 더 여유)
                yield f"data: {json.dumps(msg, ensure_ascii=False)}\n\n"
                if msg['type'] in ('result', 'error'):
                    break
                if msg['type'] in ('select', 'select_area'):
                    continue
            except queue.Empty:
                yield "data: {\"type\":\"timeout\"}\n\n"
                break

    return Response(event_stream(), mimetype='text/event-stream',
                    headers={'Cache-Control':'no-cache','X-Accel-Buffering':'no'})


@app.route('/api/kb-select', methods=['POST'])
def kb_select():
    """사용자가 다중 결과에서 선택한 항목을 Selenium 스레드로 전달"""
    body = request.json
    sid  = body.get('sid','')
    q    = _kb_progress_queues.get(sid)
    if not q:
        return jsonify({'error':'세션 없음'}), 404
    msg_type = body.get('type', 'select_result')
    q.put({'type': msg_type, 'index': body.get('index', 0),
           'area': body.get('area','')})
    return jsonify({'ok': True})


# ── 실행 ──────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    PORT = 5050
    def _open():
        import time; time.sleep(1.2)
        webbrowser.open(f'http://localhost:{PORT}')
    print("="*50)
    print("  KB부동산 시세 분석기 - Mac 웹 버전 v1")
    print(f"  http://localhost:{PORT}")
    print("  종료: Ctrl+C")
    print("="*50)
    threading.Thread(target=_open, daemon=True).start()
    app.run(host='0.0.0.0', port=PORT, debug=False)
